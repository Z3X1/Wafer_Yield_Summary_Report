"""
Wafer_Summary_Report
--------------------
Decodes all proprietary XML files in target folders
(encoding: hex string -> XOR 0xFF -> string reversal),
then generates Wafer_Summary.xlsx and Data_Summary.xlsx.

Primary save path : Z:\ToFTP  (both reports)
                    Z:\KYEC   (Wafer Summary only, additional copy)
Fallback path     : C:\KGD_data\Molex_KGD_Data  (when Z: is unavailable)
"""

import os
import re
import glob
import threading
import xml.etree.ElementTree as ET
from collections import OrderedDict
from typing import Optional

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    _HAS_DND = True
except ImportError:
    _HAS_DND = False

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Save paths ────────────────────────────────────────────────────────────────
PATH_PRIMARY  = r"Z:\ToFTP"               # Wafer Summary + Data Summary
PATH_KYEC     = r"Z:\KYEC"               # Additional copy of Wafer Summary
PATH_FALLBACK = r"C:\KGD_data\Molex_KGD_Data"  # Fallback when Z: unavailable

# ── Style constants ───────────────────────────────────────────────────────────
FILL_PASS   = PatternFill("solid", fgColor="ADD8E6")  # Light blue  - Pass die
FILL_FAIL   = PatternFill("solid", fgColor="FCD5B5")  # Orange      - Fail die (Accent 6, 60% lighter)
FILL_YAXIS  = PatternFill("solid", fgColor="D3D3D3")  # Light grey  - Y-axis / X-axis header
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")  # Light red   - Out-of-spec cell

FONT_NORMAL = Font(name="Calibri")
FONT_BOLD   = Font(name="Calibri", bold=True)

_THIN        = Side(style="thin")
BORDER_THIN  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
ALIGN_RIGHT  = Alignment(horizontal="right")
ALIGN_CENTER = Alignment(horizontal="center")

YAXIS_WIDTH = 3.5   # Column width for Y-axis label (fits 2-digit numbers)
MAP_WIDTH   = 6.0   # Column width for Wafer Map cells (fits 4-digit bin codes)


# ── XML decoding ──────────────────────────────────────────────────────────────
def decode_xml(filepath):
    # type: (str) -> Optional[str]
    """
    Decode proprietary XML encoding:
      1. Read file as hex string
      2. Convert hex to bytes
      3. XOR every byte with 0xFF
      4. Reverse the full string
      5. Extract <?xml ... </LOT_RECORD> substring
    Returns decoded XML string, or None on failure.
    """
    try:
        with open(filepath, "r", errors="ignore") as f:
            raw_hex = f.read().strip()
        raw_bytes = bytes.fromhex(raw_hex)
        decoded   = bytes(b ^ 0xFF for b in raw_bytes).decode("ascii", errors="replace")
        text      = decoded[::-1]
        start = text.find("<?xml")
        end   = text.rfind("</LOT_RECORD>")
        if start == -1 or end == -1:
            return None
        return text[start: end + len("</LOT_RECORD>")]
    except Exception:
        return None


def parse_xml(xml_str):
    # type: (str) -> Optional[dict]
    """
    Parse decoded XML string into a structured dict.
    Returns None if input is empty or XML is malformed.
    """
    if not xml_str:
        return None
    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError:
        return None

    def txt(node, tag, default=""):
        el = node.find(tag)
        return (el.text or default).strip() if el is not None else default

    # Lot-level metadata
    lot = {k: txt(root, k) for k in (
        "LOT_ID", "MODE", "PART_NUM", "PART_REV", "WO_ID", "CAL_FILE",
        "PROGRAM", "PROGRAM_REV", "MFG_PROC_ID", "MFG_PROC_REV", "OPERATOR",
        "TEST_TEMP", "TEST_FACILITY", "TESTER_ID", "START_DATE", "END_DATE",
        "TEST_CODE", "TESTER_TYPE", "SUBLOT_ID", "WO_OPCODE",
    )}

    wafer_el = root.find("WAFER_RECORD")
    if wafer_el is None:
        return None

    # Part (die) records
    parts = []
    for p in wafer_el.findall("PART_RECORD"):
        tests = OrderedDict()
        for tr in p.findall("TEST_RECORD"):
            tname = txt(tr, "TNAME")
            if tname:
                tests[tname] = {
                    "VALUE":      txt(tr, "VALUE"),
                    "LOW_LIMIT":  txt(tr, "LOW_LIMIT"),
                    "HIGH_LIMIT": txt(tr, "HIGH_LIMIT"),
                    "UNITS":      txt(tr, "UNITS"),
                    "RESULT":     txt(tr, "RESULT"),
                }
        parts.append({
            "SITE_NUM":  txt(p, "SITE_NUM"),
            "PART_ID":   txt(p, "PART_ID"),
            "RESULT":    txt(p, "RESULT"),
            "TEST_TIME": txt(p, "TEST_TIME"),
            "HARD_BIN":  txt(p, "HARD_BIN"),
            "X_COORD":   txt(p, "X_COORD"),
            "Y_COORD":   txt(p, "Y_COORD"),
            "SOFT_BIN":  txt(p, "SOFT_BIN"),
            "SOFT_DESC": txt(p, "SOFT_DESC"),
            "RETEST":    txt(p, "RETEST"),
            "tests":     tests,
        })

    return {
        "lot":   lot,
        "x_dir": txt(wafer_el, "X_DIR"),
        "y_dir": txt(wafer_el, "Y_DIR"),
        "parts": parts,
    }


def is_out_of_spec(val_str, low_str, high_str):
    # type: (str, str, str) -> bool
    """Return True if val_str exceeds HIGH_LIMIT or falls below LOW_LIMIT."""
    try:
        v = float(val_str)
    except (ValueError, TypeError):
        return False
    if high_str and high_str.strip():
        try:
            if v > float(high_str):
                return True
        except ValueError:
            pass
    if low_str and low_str.strip():
        try:
            if v < float(low_str):
                return True
        except ValueError:
            pass
    return False


def _bin_sort_key(item, bin_info):
    # type: (tuple, dict) -> tuple
    """
    Sort key for Test Summary table:
      - SOFT_BIN '1' (Pass) always first
      - All other bins sorted by count descending (highest yield impact first)
    """
    sb = item[0]
    if sb == "1":
        return (0, 0)
    return (1, -bin_info[sb]["count"])


def sanitize_filename(name):
    # type: (str) -> str
    """Strip characters not allowed in Windows filenames."""
    return re.sub(r'[\\/:*?"<>|]', "_", name)


# ── Wafer Summary report builder ──────────────────────────────────────────────
def build_wafer_summary(wafer_id_user, lot_id, part_num, records):
    # type: (str, str, str, list) -> Workbook
    """
    Build Wafer_Summary.xlsx containing:
      - Wafer Info table  (rows 1-7)
      - Test Summary table (rows 9+)
      - Wafer Map grid    (column F onwards)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Wafer_Summary"

    total  = len(records)
    passed = sum(1 for r in records if r["HARD_BIN"] == "1")
    failed = total - passed

    def bc(row, col, value=None, align=None, bold=False):
        """Write a cell with thin border and Calibri font."""
        cell = ws.cell(row, col, value)
        cell.border = BORDER_THIN
        cell.font   = FONT_BOLD if bold else FONT_NORMAL
        if align:
            cell.alignment = align
        return cell

    # ── Wafer Info section (row 1: plain title, rows 2-7: bordered data) ─────
    ws.cell(1, 1, "Wafer Info.").font = FONT_BOLD

    for i, (label, value) in enumerate([
        ("Lot ID",       lot_id),
        ("Part Number",  part_num),
        ("Wafer ID",     wafer_id_user),
        ("Total Tested", total),
        ("Passed",       passed),
        ("Failed",       failed),
    ], start=2):
        bc(i, 1, label)
        bc(i, 2, value, align=ALIGN_RIGHT)

    # ── Test Summary section (row 9: plain title, rows 10+: bordered data) ───
    ws.cell(9, 1, "Test Summary").font = FONT_BOLD
    bc(10, 1, "Failure Code")
    bc(10, 2, "Failure Name")
    bc(10, 3, "Count")
    bc(10, 4, "Rate")

    # Aggregate bin counts
    bin_info = {}
    for r in records:
        sb = r["SOFT_BIN"]
        if sb not in bin_info:
            bin_info[sb] = {"name": r["SOFT_DESC"], "count": 0}
        bin_info[sb]["count"] += 1

    for row_i, (sb, info) in enumerate(
            sorted(bin_info.items(), key=lambda x: _bin_sort_key(x, bin_info)),
            start=11):
        rate = info["count"] / total * 100 if total else 0
        bc(row_i, 1, sb)
        bc(row_i, 2, info["name"])
        bc(row_i, 3, info["count"])
        bc(row_i, 4, "{:.1f}%".format(rate))

    # ── Left panel column widths ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 15.0
    ws.column_dimensions["B"].width = 20.0
    ws.column_dimensions["C"].width = 13.0
    ws.column_dimensions["D"].width = 13.0
    ws.column_dimensions["E"].width = 13.0

    # ── Wafer Map ─────────────────────────────────────────────────────────────
    xs = [int(r["X_COORD"]) for r in records if r["X_COORD"].lstrip("-").isdigit()]
    ys = [int(r["Y_COORD"]) for r in records if r["Y_COORD"].lstrip("-").isdigit()]

    x_max = max(max(xs) if xs else 0, 30)
    y_max = max(max(ys) if ys else 0, 40)

    # Column widths: Y-axis label column narrower than die columns
    ws.column_dimensions["F"].width = YAXIS_WIDTH
    for x in range(0, x_max + 1):
        ws.column_dimensions[get_column_letter(7 + x)].width = MAP_WIDTH

    ws.cell(1, 6, "Wafer Map").font = FONT_BOLD

    # X-axis header row (row 2): grey background, centered
    c = ws.cell(2, 6, -1)
    c.fill = FILL_YAXIS; c.alignment = ALIGN_CENTER; c.font = FONT_NORMAL
    for x in range(0, x_max + 1):
        c = ws.cell(2, 7 + x, x)
        c.fill = FILL_YAXIS; c.alignment = ALIGN_CENTER; c.font = FONT_NORMAL

    # Y-axis labels (column F): grey background, centered
    for y in range(0, y_max + 1):
        c = ws.cell(y + 3, 6, y)
        c.fill = FILL_YAXIS; c.alignment = ALIGN_CENTER; c.font = FONT_NORMAL

    # Apply grid borders to all cells within the map range (including empty)
    for y in range(0, y_max + 1):
        for x in range(0, x_max + 1):
            ws.cell(y + 3, x + 7).border = BORDER_THIN

    # Fill die cells with pass/fail color and SOFT_BIN code
    for r in records:
        try:
            x = int(r["X_COORD"])
            y = int(r["Y_COORD"])
        except ValueError:
            continue
        cell = ws.cell(y + 3, x + 7, r["SOFT_BIN"])
        cell.fill      = FILL_PASS if r["HARD_BIN"] == "1" else FILL_FAIL
        cell.alignment = ALIGN_CENTER
        cell.font      = FONT_NORMAL

    return wb


# ── Data Summary report builder ───────────────────────────────────────────────
FIXED_HEADERS = [
    "LOT_ID", "MODE", "PART_NUM", "PART_REV", "WO_ID", "CAL_FILE",
    "PROGRAM", "PROGRAM_REV", "MFG_PROC_ID", "MFG_PROC_REV", "OPERATOR",
    "TEST_TEMP", "TEST_FACILITY", "TESTER_ID", "START_DATE", "END_DATE",
    "TEST_CODE", "TESTER_TYPE", "SUBLOT_ID", "WO_OPCODE",
    "WAFER_ID",   # user-supplied Wafer ID
    "X_DIR", "Y_DIR", "SITE_NUM", "PART_ID", "RESULT", "TEST_TIME",
    "HARD_BIN", "X_COORD", "Y_COORD", "SOFT_BIN", "SOFT_DESC", "RETEST",
]


def build_data_summary(wafer_id_user, all_records, tnames):
    # type: (str, list, list) -> Workbook
    """
    Build Data_Summary.xlsx containing one row per die with:
      - 33 fixed metadata columns
      - One column per test parameter (VALUE only; out-of-spec cells red)
    Test parameter columns with no valid numeric values across all records
    are excluded (e.g. vision-only inspection steps).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Header row
    for col, h in enumerate(FIXED_HEADERS + tnames, 1):
        ws.cell(1, col, h).font = FONT_BOLD

    # Data rows
    for row_i, rec in enumerate(all_records, start=2):
        lot  = rec["lot"]
        part = rec["part"]

        fixed_vals = [
            lot.get("LOT_ID", ""),        lot.get("MODE", ""),
            lot.get("PART_NUM", ""),       lot.get("PART_REV", ""),
            lot.get("WO_ID", ""),          lot.get("CAL_FILE", ""),
            lot.get("PROGRAM", ""),        lot.get("PROGRAM_REV", ""),
            lot.get("MFG_PROC_ID", ""),    lot.get("MFG_PROC_REV", ""),
            lot.get("OPERATOR", ""),       lot.get("TEST_TEMP", ""),
            lot.get("TEST_FACILITY", ""),  lot.get("TESTER_ID", ""),
            lot.get("START_DATE", ""),     lot.get("END_DATE", ""),
            lot.get("TEST_CODE", ""),      lot.get("TESTER_TYPE", ""),
            lot.get("SUBLOT_ID", ""),      lot.get("WO_OPCODE", ""),
            wafer_id_user,
            rec.get("x_dir", ""),          rec.get("y_dir", ""),
            part.get("SITE_NUM", ""),      part.get("PART_ID", ""),
            part.get("RESULT", ""),        part.get("TEST_TIME", ""),
            part.get("HARD_BIN", ""),      part.get("X_COORD", ""),
            part.get("Y_COORD", ""),       part.get("SOFT_BIN", ""),
            part.get("SOFT_DESC", ""),     part.get("RETEST", ""),
        ]
        for col, v in enumerate(fixed_vals, 1):
            ws.cell(row_i, col, v).font = FONT_NORMAL

        # Test parameter columns
        tests = part.get("tests", {})
        base  = len(FIXED_HEADERS) + 1
        for col_off, tname in enumerate(tnames, base):
            t    = tests.get(tname, {})
            vstr = t.get("VALUE", "")
            try:
                val = float(vstr)
            except (ValueError, TypeError):
                val = vstr
            cell = ws.cell(row_i, col_off, val)
            cell.font = FONT_NORMAL
            if is_out_of_spec(vstr, t.get("LOW_LIMIT", ""), t.get("HIGH_LIMIT", "")):
                cell.fill = FILL_RED  # Highlight out-of-spec values in red

    return wb


# ── File save helpers ─────────────────────────────────────────────────────────
def _try_save(wb, path, filename):
    # type: (Workbook, str, str) -> bool
    """Attempt to save workbook to path/filename. Returns True on success."""
    try:
        os.makedirs(path, exist_ok=True)
        wb.save(os.path.join(path, filename))
        return True
    except Exception:
        return False


def save_reports(wafer_wb, data_wb, wafer_fn, data_fn):
    # type: (Workbook, Workbook, str, str) -> tuple
    """
    Save both reports to primary paths (Z:\\ToFTP and Z:\\KYEC).
    If any primary path fails, fall back to C:\\KGD_data\\Molex_KGD_Data.
    Returns (success: bool, message: str).
    """
    ok1 = _try_save(wafer_wb, PATH_PRIMARY, wafer_fn)
    ok2 = _try_save(data_wb,  PATH_PRIMARY, data_fn)
    ok3 = _try_save(wafer_wb, PATH_KYEC,    wafer_fn)

    if ok1 and ok2 and ok3:
        return True, (
            "Reports saved successfully.\n"
            "  {0}\\{1}\n"
            "  {0}\\{2}\n"
            "  {3}\\{1}"
        ).format(PATH_PRIMARY, wafer_fn, data_fn, PATH_KYEC)

    # Fallback path
    fb1 = _try_save(wafer_wb, PATH_FALLBACK, wafer_fn)
    fb2 = _try_save(data_wb,  PATH_FALLBACK, data_fn)
    if fb1 and fb2:
        return True, (
            "Warning: Z:\\ unavailable. Reports saved to fallback path.\n"
            "  {0}\\{1}\n"
            "  {0}\\{2}"
        ).format(PATH_FALLBACK, wafer_fn, data_fn)

    return False, "Error: All save paths failed. Please check drive connections."


# ── Main processing pipeline ──────────────────────────────────────────────────
def process(wafer_id_user, folder_list, progress_cb, status_cb):
    # type: (str, list, object, object) -> tuple
    """
    Full pipeline:
      1. Discover all XML files under given folders (recursive)
      2. Decode and parse each XML
      3. Validate single LOT_ID across all files
      4. Build Wafer Summary and Data Summary workbooks
      5. Save to configured paths
    Returns (success: bool, message: str).
    """

    # Step 1: Collect XML files
    xml_files = []
    for folder in folder_list:
        xml_files += glob.glob(
            os.path.join(folder, "**", "*.xml"), recursive=True)
    xml_files = sorted(set(xml_files))

    if not xml_files:
        return False, "No XML files found in the specified folders."

    status_cb("Found {} XML file(s). Parsing...".format(len(xml_files)))

    parsed        = []
    lot_ids       = set()
    part_nums     = set()
    all_tnames    = OrderedDict()
    tname_has_val = {}  # track which test params have at least one numeric value

    # Step 2: Decode and parse
    for i, fp in enumerate(xml_files):
        progress_cb((i + 1) / len(xml_files) * 0.5)
        data = parse_xml(decode_xml(fp))
        if data is None:
            continue
        lot_ids.add(data["lot"]["LOT_ID"])
        part_nums.add(data["lot"]["PART_NUM"])
        for part in data["parts"]:
            for tname, tdata in part["tests"].items():
                all_tnames.setdefault(tname, None)
                if tdata.get("VALUE", "").strip():
                    tname_has_val[tname] = True  # has at least one real value
        parsed.append(data)

    # Exclude test columns that have no numeric values across all records
    all_tnames = OrderedDict(
        (k, v) for k, v in all_tnames.items() if tname_has_val.get(k, False)
    )

    # Step 3: Validate LOT_ID uniqueness
    if len(lot_ids) > 1:
        return False, (
            "Error: Multiple LOT_IDs detected in the provided XML files:\n"
            + "\n".join("  - {}".format(v) for v in sorted(lot_ids))
            + "\nAll XML files must belong to the same lot."
        )
    if not parsed:
        return False, "Error: Failed to parse any XML files. Please verify file format."

    lot_id   = next(iter(lot_ids))
    part_num = next(iter(part_nums)) if part_nums else ""
    tnames   = list(all_tnames)

    # Flatten records
    flat_records = []  # die-level only (for Wafer Summary)
    full_records = []  # full lot + die context (for Data Summary)
    for data in parsed:
        for part in data["parts"]:
            flat_records.append(part)
            full_records.append({
                "lot":   data["lot"],
                "x_dir": data["x_dir"],
                "y_dir": data["y_dir"],
                "part":  part,
            })

    # Step 4: Build reports
    status_cb("Building Wafer Summary...")
    progress_cb(0.6)
    wafer_wb = build_wafer_summary(wafer_id_user, lot_id, part_num, flat_records)

    status_cb("Building Data Summary...")
    progress_cb(0.8)
    data_wb = build_data_summary(wafer_id_user, full_records, tnames)

    # Step 5: Save
    status_cb("Saving files...")
    safe_wid = sanitize_filename(wafer_id_user)
    base     = "{}_{}".format(safe_wid, part_num)
    wafer_fn = "{}_Wafer_Summary.xlsx".format(base)
    data_fn  = "{}_Data_Summary.xlsx".format(base)

    ok, msg = save_reports(wafer_wb, data_wb, wafer_fn, data_fn)
    progress_cb(1.0)
    return ok, msg


# ── GUI ───────────────────────────────────────────────────────────────────────
class App(object):
    def __init__(self, root):
        self.root     = root
        self._running = False   # prevents duplicate generate calls
        self._btn_gen = None

        root.title("Wafer_Summary_Report")
        root.resizable(False, False)
        root.configure(bg="#f0f0f0")

        # Wafer ID input row
        frm_top = tk.Frame(root, bg="#f0f0f0")
        frm_top.pack(padx=20, pady=(20, 5), fill="x")
        tk.Label(frm_top, text="Wafer ID:",
                 bg="#f0f0f0", font=("Arial", 11)).pack(side="left")
        self.var_wid = tk.StringVar()
        tk.Entry(frm_top, textvariable=self.var_wid,
                 font=("Arial", 11), width=35).pack(side="left", padx=(8, 0))

        # Instruction label
        tk.Label(root,
                 text="Drag folders into the area below (multiple allowed, must contain XML files)",
                 bg="#f0f0f0", font=("Arial", 10)).pack(anchor="w", padx=20)

        # Folder list with scrollbar
        frm_list = tk.Frame(root, bg="#f0f0f0")
        frm_list.pack(padx=20, pady=5, fill="both", expand=True)
        self.listbox = tk.Listbox(
            frm_list, selectmode=tk.EXTENDED,
            font=("Arial", 10), height=8, width=65, activestyle="dotbox")
        sb = tk.Scrollbar(frm_list, orient="vertical",
                          command=self.listbox.yview)
        self.listbox.config(yscrollcommand=sb.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        # Register drag-and-drop target if available
        if _HAS_DND:
            self.listbox.drop_target_register(DND_FILES)
            self.listbox.dnd_bind("<<Drop>>", self._on_drop)

        # Folder management buttons
        frm_btns = tk.Frame(root, bg="#f0f0f0")
        frm_btns.pack(padx=20, pady=5, fill="x")
        bkw = {"font": ("Arial", 10), "width": 12, "relief": "raised", "bd": 2}
        tk.Button(frm_btns, text="Add Folder",    bg="#4CAF50", fg="white",
                  command=self._add_folder,        **bkw).pack(side="left", padx=(0, 6))
        tk.Button(frm_btns, text="Remove Selected", bg="#795548", fg="white",
                  command=self._remove_selected,   **bkw).pack(side="left", padx=(0, 6))
        tk.Button(frm_btns, text="Clear All",     bg="#9E9E9E", fg="white",
                  command=self._clear_list,        **bkw).pack(side="left")

        # Progress bar
        tk.Label(root, text="Progress:", bg="#f0f0f0",
                 font=("Arial", 10)).pack(anchor="w", padx=20, pady=(10, 0))
        self.progress = ttk.Progressbar(root, length=520, mode="determinate")
        self.progress.pack(padx=20, pady=(0, 5))
        self.lbl_status = tk.Label(root, text="", bg="#f0f0f0",
                                   font=("Arial", 10), wraplength=520)
        self.lbl_status.pack(padx=20, pady=(0, 8))

        # Action buttons
        frm_bot = tk.Frame(root, bg="#f0f0f0")
        frm_bot.pack(padx=20, pady=(0, 20), fill="x")
        self._btn_gen = tk.Button(
            frm_bot, text="Generate Report",
            bg="#3F51B5", fg="white",
            font=("Arial", 12, "bold"), width=16, command=self._generate)
        self._btn_gen.pack(side="left", padx=(0, 20))
        tk.Button(frm_bot, text="Exit", bg="#F44336", fg="white",
                  font=("Arial", 12, "bold"), width=12,
                  command=root.destroy).pack(side="left")

    # ── Event handlers ────────────────────────────────────────────────────────
    def _on_drop(self, event):
        """Handle folder drag-and-drop into the listbox."""
        paths = self.root.tk.splitlist(event.data)
        for p in paths:
            p = p.strip("{}")
            if os.path.isdir(p) and p not in self.listbox.get(0, "end"):
                self.listbox.insert("end", p)

    def _add_folder(self):
        """Open a folder browser dialog and add selection to the list."""
        folder = filedialog.askdirectory(title="Select folder containing XML files")
        if folder and folder not in self.listbox.get(0, "end"):
            self.listbox.insert("end", folder)

    def _remove_selected(self):
        """Remove currently selected entries from the listbox."""
        for i in reversed(self.listbox.curselection()):
            self.listbox.delete(i)

    def _clear_list(self):
        """Clear all entries from the listbox."""
        self.listbox.delete(0, "end")

    def _set_status(self, msg):
        self.root.after(0, lambda: self.lbl_status.config(text=msg))

    def _set_progress(self, val):
        self.root.after(0, lambda: self.progress.config(value=val * 100))

    def _set_busy(self, busy):
        """Lock/unlock the Generate button to prevent duplicate execution."""
        state = "disabled" if busy else "normal"
        self.root.after(0, lambda: self._btn_gen.config(state=state))
        self._running = busy

    def _generate(self):
        """Validate inputs and launch the processing pipeline in a background thread."""
        if self._running:
            return

        wafer_id = self.var_wid.get().strip()
        if not wafer_id:
            messagebox.showwarning(
                "Missing Wafer ID",
                "Please enter a Wafer ID before generating the report.")
            return

        folders = list(self.listbox.get(0, "end"))
        if not folders:
            messagebox.showwarning(
                "No Folders",
                "Please add at least one folder containing XML files.")
            return

        self.progress.config(value=0)
        self._set_status("Processing...")
        self._set_busy(True)

        def run():
            ok, msg = process(
                wafer_id, folders, self._set_progress, self._set_status)
            def finish():
                self._set_busy(False)
                if ok:
                    self._set_status(msg)
                else:
                    messagebox.showerror("Error", msg)
                    self._set_status("Error occurred. See dialog for details.")
            self.root.after(0, finish)

        threading.Thread(target=run, daemon=True).start()


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if _HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()

    # Center window on screen
    w, h = 600, 490
    root.update_idletasks()
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    root.geometry("{0}x{1}+{2}+{3}".format(
        w, h, (sw - w) // 2, (sh - h) // 2))

    App(root)
    root.mainloop()
